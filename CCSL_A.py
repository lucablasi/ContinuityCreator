def ccsl_a(sublist, fr, m_title):
    """
    Create CCSL excel file from sublist.
    :param sublist: srt subtitle object list with dialogue, caption, sub list
    :param fr: framerate
    :param m_title: movie title
    :return:
    """

    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.worksheet.dimensions import SheetFormatProperties
    import srt
    from timecoder import tc_f
    from timecoder import tc_dur

    # Workbook setup
    wb = openpyxl.load_workbook('CCSL Template.xlsx')
    ws = wb.active
    ws.title = m_title
    ws.sheet_format = SheetFormatProperties(defaultRowHeight=13)
    ws.cell(1, 1, m_title + ' - Combined Continuity & Spotting List')
    ws.print_title_rows = '3:4'
    thin = Side(border_style='thin')

    # Variable initialization
    r = 5  # Starting row
    r_left = r  # Content left row tracker
    r_right = r  # Content right row tracker
    take_list = []  # Take row list
    index = 1  # Subtitle index

    # Content fill!
    for i in range(len(sublist)):
        print(str(i+1) + ' of ' + str(len(sublist)))
        if 'scene' in sublist[i].proprietary:  # Scenes
            # Timestamps
            ws.cell(r_left, 1, tc_f(srt.timedelta_to_srt_timestamp(sublist[i].start), fr))
            ws.cell(r_left, 2, tc_dur(sublist[i], fr))
            ws.cell(r_left, 3, tc_f(srt.timedelta_to_srt_timestamp(sublist[i].end), fr))
            ws.cell(r_left, 1).font = Font(bold=True)
            ws.cell(r_left, 2).font = Font(bold=True)
            ws.cell(r_left, 3).font = Font(bold=True)
            ws.cell(r_left, 1).alignment = Alignment(horizontal='center')
            ws.cell(r_left, 2).alignment = Alignment(horizontal='center')
            ws.cell(r_left, 3).alignment = Alignment(horizontal='center')
            take_list.append(r_left)

            ws.cell(r_left, 1).border = Border(left=thin, bottom=thin, top=thin)
            ws.cell(r_left, 2).border = Border(bottom=thin, top=thin)
            ws.cell(r_left, 3).border = Border(bottom=thin, top=thin, right=thin)
            ws.cell(r_left, 4).border = Border(bottom=thin, top=thin)
            ws.cell(r_left, 5).border = Border(bottom=thin, top=thin)
            ws.cell(r_left, 6).border = Border(bottom=thin, top=thin)
            ws.cell(r_left, 7).border = Border(bottom=thin, top=thin)
            ws.cell(r_left, 8).border = Border(bottom=thin, top=thin)
            ws.cell(r_left, 9).border = Border(left=thin, bottom=thin, right=thin, top=thin)

            for col in range(1, 10):
                ws.cell(r_left, col).fill = PatternFill("solid", fgColor="DDDDDD")

            r_left += 1
            ws.cell(r_left, 1).border = Border(left=thin)
            ws.cell(r_left, 3).border = Border(right=thin)
            ws.cell(r_left, 9).border = Border(left=thin, right=thin)

            r_right = r_left

            # Content
            for j in range(len(sublist[i].content)):
                ws.cell(r_left, 1, sublist[i].content[j])
                ws.cell(r_left, 1).border = Border(left=thin)
                ws.cell(r_left, 3).border = Border(right=thin)
                ws.cell(r_left, 9).border = Border(left=thin, right=thin)
                r_left += 1
            ws.cell(r_left, 1).border = Border(left=thin)
            ws.cell(r_left, 3).border = Border(right=thin)
            ws.cell(r_left, 9).border = Border(left=thin, right=thin)
            r_left += 1

        else:  # Subs
            # Index, note, and timestamps
            ws.cell(r_right, 4, index)
            index += 1
            ws.cell(r_right, 4).alignment = Alignment(horizontal='center')

            if 'ital' in sublist[i].proprietary:
                ws.cell(r_right, 5, 'ital')
                ws.cell(r_right, 5).alignment = Alignment(horizontal='center')
                ws.cell(r_right, 5).font = Font(italic=True)
                ws.cell(r_right, 9).font = Font(italic=True)

            ws.cell(r_right, 6, tc_f(srt.timedelta_to_srt_timestamp(sublist[i].start), fr))
            ws.cell(r_right, 6).alignment = Alignment(horizontal='center')

            ws.cell(r_right, 7, tc_dur(sublist[i], fr))
            ws.cell(r_right, 7).alignment = Alignment(horizontal='center')

            ws.cell(r_right, 8, tc_f(srt.timedelta_to_srt_timestamp(sublist[i].end), fr))
            ws.cell(r_right, 8).alignment = Alignment(horizontal='center')

            # Content
            for j in range(len(sublist[i].content)):
                ws.cell(r_right, 9, sublist[i].content[j])
                ws.cell(r_right, 1).border = Border(left=thin)
                ws.cell(r_right, 3).border = Border(right=thin)
                ws.cell(r_right, 9).border = Border(left=thin, right=thin)
                r_right += 1
                ws.cell(r_right, 1).border = Border(left=thin)
                ws.cell(r_right, 3).border = Border(right=thin)
                ws.cell(r_right, 9).border = Border(left=thin, right=thin)
            r_right += 1
            ws.cell(r_right, 1).border = Border(left=thin)
            ws.cell(r_right, 3).border = Border(right=thin)
            ws.cell(r_right, 9).border = Border(left=thin, right=thin)

        # Start new take after lowest row entered
        if r_right > r_left:
            r_left = r_right

    # End bottom lines
    if r_right > r_left:
        r_end = r_right
    else:
        r_end = r_left
    for i in range(1, 10):
        ws.cell(r_end, i).border = Border(top=thin)

    # Add rows to fit take start at page breaks
    brk = 79  # First page break row
    comp = 0  # Compensation for added rows
    for i in range(len(take_list)):
        if i + 1 < len(take_list):
            if take_list[i + 1] - take_list[i] < 77:
                if take_list[i] <= brk < take_list[i + 1]:
                    n = brk - take_list[i] + 1
                    ws.insert_rows(take_list[i]+comp, n)
                    for j in range(take_list[i]+comp, brk+comp):
                            ws.cell(j, 1).border = Border(left=thin)
                            ws.cell(j, 3).border = Border(right=thin)
                            ws.cell(j, 9).border = Border(left=thin, right=thin)
                    ws.cell(brk + comp, 1).border = Border(left=thin, bottom=thin)
                    ws.cell(brk + comp, 2).border = Border(bottom=thin)
                    ws.cell(brk + comp, 3).border = Border(right=thin, bottom=thin)
                    ws.cell(brk + comp, 4).border = Border(bottom=thin)
                    ws.cell(brk + comp, 5).border = Border(bottom=thin)
                    ws.cell(brk + comp, 6).border = Border(bottom=thin)
                    ws.cell(brk + comp, 7).border = Border(bottom=thin)
                    ws.cell(brk + comp, 8).border = Border(bottom=thin)
                    ws.cell(brk + comp, 9).border = Border(left=thin, bottom=thin, right=thin)

                    comp += n
                    brk += 77 - n  # 77 rows per page after first page
            else:  # Case where one take is longer than whole page
                ws.cell(brk + comp, 1).border = Border(left=thin, bottom=thin)
                ws.cell(brk + comp, 2).border = Border(bottom=thin)
                ws.cell(brk + comp, 3).border = Border(right=thin, bottom=thin)
                ws.cell(brk + comp, 4).border = Border(bottom=thin)
                ws.cell(brk + comp, 5).border = Border(bottom=thin)
                ws.cell(brk + comp, 6).border = Border(bottom=thin)
                ws.cell(brk + comp, 7).border = Border(bottom=thin)
                ws.cell(brk + comp, 8).border = Border(bottom=thin)
                ws.cell(brk + comp, 9).border = Border(left=thin, bottom=thin, right=thin)
                brk += 77

    print('Saving!')
    wb.save(m_title + ' CCSL.xlsx')
