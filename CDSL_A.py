def cdsl_a(sclist, sublist, fr, m_title):
    """
    Create CDSL excel file from sublists. In particular, dialogue, caption,
    and subtitle sublists are fitted within their corresponding scene from
    the scene sublist
    :param sclist: scene sublist
    :param sublist: srt subtitle object list with dialogue, caption, sub list
    :param fr: framerate
    :param m_title: movie title
    :return:
    """

    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.worksheet.dimensions import SheetFormatProperties
    import srt
    from timecoder import tc_f
    from timecoder import tc_dur

    # Workbook setup
    wb = openpyxl.load_workbook('CDSL Template.xlsx')
    ws = wb.active
    ws.title = m_title
    ws.sheet_format = SheetFormatProperties(defaultRowHeight=13)
    ws.cell(1, 1, m_title + ' - Continuity Dialogue Spotting List')
    ws.print_title_rows = '3:4'
    thin = Side(border_style='thin')

    # Variable initialization
    r = 5           # Starting row
    r_left = r      # Content left row tracker
    r_right = r     # Content right row tracker
    take = 1        # Take num index
    take_list = []  # Take row list
    index = 1       # Subtitle index
    j = 0           # Content counter

    # Content fill!
    for i in range(len(sclist)):
        # Take timestamp
        ws.merge_cells(start_row=r_left, start_column=1, end_row=r_left, end_column=3)
        take_content = 'Take ' + str(take) + ' - ' + tc_f(srt.timedelta_to_srt_timestamp(sclist[i].start), fr)
        ws.cell(r_left, 1, take_content)
        ws.cell(r_left, 1).font = Font(bold=True)
        take += 1

        take_list.append(r_left)

        ws.cell(r_left, 1).border = Border(left=thin, bottom=thin, top=thin)
        ws.cell(r_left, 2).border = Border(bottom=thin, top=thin)
        ws.cell(r_left, 3).border = Border(bottom=thin, top=thin, right=thin)
        ws.cell(r_left, 4).border = Border(bottom=thin, top=thin)
        ws.cell(r_left, 5).border = Border(bottom=thin, top=thin)
        ws.cell(r_left, 6).border = Border(bottom=thin, top=thin)
        ws.cell(r_left, 7).border = Border(bottom=thin, top=thin)
        ws.cell(r_left, 8).border = Border(bottom=thin, top=thin)
        ws.cell(r_left, 9).border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for col in range(1, 10):
            ws.cell(r_left, col).fill = PatternFill("solid", fgColor="DDDDDD")

        r_line = r_left
        r_left += 2
        r_right = r_left

        # Dialogue, caption and title
        while sublist[j].start < sclist[i].end:  # Fit subs within takes
            if 'dialogue' in sublist[j].proprietary or 'caption' in sublist[j].proprietary:
                for k in range(len(sublist[j].content)):
                    ws.merge_cells(start_row=r_left, start_column=1, end_row=r_left, end_column=3)
                    ws.cell(r_left, 1, sublist[j].content[k])
                    r_left += 1
                r_left += 1

                j += 1

            elif 'title' in sublist[j].proprietary:
                # Index, note, and timestamps
                ws.cell(r_right, 4, index)
                index += 1
                ws.cell(r_right, 4).alignment = Alignment(horizontal='center')

                if 'ital' in sublist[j].proprietary:
                    ws.cell(r_right, 5, 'ital')
                    ws.cell(r_right, 5).alignment = Alignment(horizontal='center')
                    ws.cell(r_right, 5).font = Font(italic=True)
                    ws.cell(r_right, 9).font = Font(italic=True)

                ws.cell(r_right, 6, tc_f(srt.timedelta_to_srt_timestamp(sublist[j].start), fr))
                ws.cell(r_right, 6).alignment = Alignment(horizontal='center')

                ws.cell(r_right, 7, tc_dur(sublist[j], fr))
                ws.cell(r_right, 7).alignment = Alignment(horizontal='center')

                ws.cell(r_right, 8, tc_f(srt.timedelta_to_srt_timestamp(sublist[j].end), fr))
                ws.cell(r_right, 8).alignment = Alignment(horizontal='center')

                # Content
                for k in range(len(sublist[j].content)):
                        ws.cell(r_right, 9, sublist[j].content[k])
                        if r_right == r_line:
                            ws.cell(r_right, 9).border = Border(right=thin, top=thin)
                        else:
                            ws.cell(r_right, 9).border = Border(right=thin)
                            ws.cell(r_right, 4).border = Border(left=thin)
                        r_right += 1
                r_right += 1

                j += 1

            if j >= len(sublist):
                break

        # Start new take after lowest row entered
        if r_right > r_left:
            r_end = r_right
            r_left = r_right
        else:
            r_end = r_left

        # In between takes vertical lines
        for row in range(r_line+1, r_end+1):
            ws.cell(row, 1).border = Border(left=thin)
            ws.cell(row, 3).border = Border(right=thin)
            ws.cell(row, 9).border = Border(left=thin, right=thin)

    # End bottom lines
    if r_right > r_left:
        r_end = r_right
    else:
        r_end = r_left
    for i in range(1, 10):
        ws.cell(r_end, i).border = Border(top=thin)

    # Add rows to fit take start at page breaks
    brk = 79  # First page break row
    comp = 0  # Compensation for added rows
    for i in range(len(take_list)):
        if i + 1 < len(take_list):                      # Case for last take
            if take_list[i + 1] - take_list[i] < 77:    # Case where one take is longer than whole page
                if take_list[i] <= brk < take_list[i + 1]:
                    n = brk - take_list[i] + 1
                    ws.insert_rows(take_list[i]+comp, n)
                    for j in range(take_list[i]+comp, brk+comp):
                            ws.cell(j, 1).border = Border(left=thin)
                            ws.cell(j, 3).border = Border(right=thin)
                            ws.cell(j, 9).border = Border(left=thin, right=thin)

                    ws.cell(brk + comp, 1).border = Border(left=thin, bottom=thin)
                    ws.cell(brk + comp, 2).border = Border(bottom=thin)
                    ws.cell(brk + comp, 3).border = Border(right=thin, bottom=thin)
                    ws.cell(brk + comp, 4).border = Border(bottom=thin)
                    ws.cell(brk + comp, 5).border = Border(bottom=thin)
                    ws.cell(brk + comp, 6).border = Border(bottom=thin)
                    ws.cell(brk + comp, 7).border = Border(bottom=thin)
                    ws.cell(brk + comp, 8).border = Border(bottom=thin)
                    ws.cell(brk + comp, 9).border = Border(left=thin, bottom=thin, right=thin)

                    comp += n
                    brk += 77 - n  # 77 rows per page after first page
            else:
                ws.cell(brk + comp, 1).border = Border(left=thin, bottom=thin)
                ws.cell(brk + comp, 2).border = Border(bottom=thin)
                ws.cell(brk + comp, 3).border = Border(right=thin, bottom=thin)
                ws.cell(brk + comp, 4).border = Border(bottom=thin)
                ws.cell(brk + comp, 5).border = Border(bottom=thin)
                ws.cell(brk + comp, 6).border = Border(bottom=thin)
                ws.cell(brk + comp, 7).border = Border(bottom=thin)
                ws.cell(brk + comp, 8).border = Border(bottom=thin)
                ws.cell(brk + comp, 9).border = Border(left=thin, bottom=thin, right=thin)
                brk += 77
        else:
            ws.cell(brk + comp, 1).border = Border(left=thin, bottom=thin)
            ws.cell(brk + comp, 2).border = Border(bottom=thin)
            ws.cell(brk + comp, 3).border = Border(right=thin, bottom=thin)
            ws.cell(brk + comp, 4).border = Border(bottom=thin)
            ws.cell(brk + comp, 5).border = Border(bottom=thin)
            ws.cell(brk + comp, 6).border = Border(bottom=thin)
            ws.cell(brk + comp, 7).border = Border(bottom=thin)
            ws.cell(brk + comp, 8).border = Border(bottom=thin)
            ws.cell(brk + comp, 9).border = Border(left=thin, bottom=thin, right=thin)

    wb.save(m_title + ' CDSL.xlsx')
